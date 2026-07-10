import json
import os
import re
import secrets
import time
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

import bcrypt
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DATA_PATH = ROOT / "data.js"
EXCEL_PATH = ROOT / "sample_data_cleaned.xlsx"
USERS_PATH = ROOT / "users.json"
SESSION_SECONDS = 60 * 60 * 8
RESET_CODE_SECONDS = 10 * 60
RESET_TOKEN_SECONDS = 15 * 60
SESSIONS = {}
PASSWORD_RESETS = {}
ROOT_ADMIN_EMAIL = "alharbi.moh2003@example-company.com"
WORKBOOK_SHEETS = {
    "Dashboard",
    "Departments",
    "Employees",
    "Projects",
    "Tasks",
    "Meetings",
    "Weekly Updates",
    "Activity Log",
    "Lists",
}


def load_env():
    env_path = ROOT / ".env"
    if not env_path.exists():
        return

    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def employee_records():
    workbook = load_workbook_data()
    return workbook.get("Employees", [])


def load_workbook_data():
    match = re.search(r"window\.WORKBOOK_DATA\s*=\s*(\{.*\});?\s*$", DATA_PATH.read_text(encoding="utf-8"))
    if not match:
        return {}
    return json.loads(match.group(1))


def save_workbook_data(workbook):
    body = json.dumps(workbook, separators=(",", ":"), ensure_ascii=False)
    DATA_PATH.write_text(f"window.WORKBOOK_DATA = {body};", encoding="utf-8")


def save_workbook_to_excel(workbook_data):
    if not EXCEL_PATH.exists():
        return

    workbook = load_workbook(EXCEL_PATH)
    for sheet_name, rows in workbook_data.items():
        if sheet_name not in workbook.sheetnames or not isinstance(rows, list):
            continue
        sheet = workbook[sheet_name]
        headers = excel_headers(sheet)
        if sheet.max_row > 1:
            sheet.delete_rows(2, sheet.max_row - 1)
        for row in rows:
            if isinstance(row, dict):
                sheet.append([row.get(header, "") for header in headers])

    workbook.save(EXCEL_PATH)
    workbook.close()


def replace_workbook_data(workbook):
    cleaned = {}
    for sheet_name, rows in workbook.items():
        if sheet_name not in WORKBOOK_SHEETS or not isinstance(rows, list):
            continue
        cleaned[sheet_name] = [row for row in rows if isinstance(row, dict)]

    if "Employees" not in cleaned:
        return False, "Employees sheet is required."

    save_workbook_data(cleaned)
    save_workbook_to_excel(cleaned)
    return True, ""


def employee_emails():
    emails = set()
    for employee in employee_records():
        email = employee.get("Email")
        if email:
            emails.add(email.strip().lower())
    return emails


def employee_by_email(email):
    normalized = email.strip().lower()
    for employee in employee_records():
        if str(employee.get("Email", "")).strip().lower() == normalized:
            return employee
    return {}


def company_domain():
    domains = sorted({email.rsplit("@", 1)[1] for email in employee_emails() if "@" in email})
    return domains[0] if domains else "example-company.com"


def next_employee_id(employees):
    highest = 1000
    for employee in employees:
        match = re.fullmatch(r"E(\d+)", str(employee.get("Employee ID", "")))
        if match:
            highest = max(highest, int(match.group(1)))
    return f"E{highest + 1}"


def new_employee_record(first_name, last_name, email, employee_id):
    normalized = email.strip().lower()
    return {
        "Employee ID": employee_id,
        "Employee Name": f"{first_name} {last_name}".strip(),
        "Employee Name First Name": first_name,
        "Employee Name Last Name": last_name,
        "Email": normalized,
        "Department ID": "",
        "Department": "Pending Assignment",
        "Job Title": "New User",
        "Level": "New User",
        "Manager": "",
        "Manager First Name": "",
        "Manager Last Name": "",
        "Location": "",
        "Hire Date": time.strftime("%Y-%m-%d"),
        "Employment Status": "Active",
    }


def add_employee_to_dashboard_data(first_name, last_name, email):
    normalized = email.strip().lower()
    workbook = load_workbook_data()
    employees = workbook.setdefault("Employees", [])
    for employee in employees:
        if str(employee.get("Email", "")).strip().lower() == normalized:
            return employee

    employee = new_employee_record(first_name, last_name, normalized, next_employee_id(employees))
    employees.append(employee)
    save_workbook_data(workbook)
    return employee


def excel_headers(sheet):
    return [cell.value for cell in sheet[1]]


def add_employee_to_excel_data(employee):
    if not EXCEL_PATH.exists():
        return

    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook["Employees"]
    headers = excel_headers(sheet)
    email_column = headers.index("Email") + 1
    normalized = employee["Email"].strip().lower()

    for row_index in range(2, sheet.max_row + 1):
        if str(sheet.cell(row_index, email_column).value or "").strip().lower() == normalized:
            workbook.close()
            return

    sheet.append([employee.get(header, "") for header in headers])
    workbook.save(EXCEL_PATH)
    workbook.close()


def add_employee_record(first_name, last_name, email):
    employee = add_employee_to_dashboard_data(first_name, last_name, email)
    add_employee_to_excel_data(employee)
    return employee


def create_dashboard_employee(first_name, last_name, email):
    first_name = first_name.strip()
    last_name = last_name.strip()
    normalized = email.strip().lower()
    domain = company_domain()

    if not first_name or not last_name or not normalized:
        return False, "First name, last name, and company email are required.", {}
    if not normalized.endswith(f"@{domain}"):
        return False, f"Use your @{domain} company email address.", {}
    if normalized in employee_emails():
        return False, "This user is already in the dashboard.", {}

    employee = add_employee_record(first_name, last_name, normalized)
    users = load_users()
    if normalized in users:
        users[normalized]["approval_status"] = "approved"
        users[normalized]["approved_at"] = int(time.time())
        users[normalized]["approved_by"] = "admin-add-user"
        save_users(users)
    return True, "", employee


def sync_registered_users_to_employees():
    domain = company_domain()
    for email, user in load_users().items():
        if user.get("approval_status") == "pending":
            continue
        if user.get("approval_status") == "rejected":
            continue
        if "@" not in email or not email.endswith(f"@{domain}"):
            continue
        first_name = user.get("first_name") or email.split("@", 1)[0].split(".", 1)[0].title()
        last_name = user.get("last_name") or ""
        add_employee_record(first_name, last_name, email)


def remove_employee_from_dashboard_data(email):
    normalized = email.strip().lower()
    workbook = load_workbook_data()
    employees = workbook.get("Employees", [])
    kept = [employee for employee in employees if str(employee.get("Email", "")).strip().lower() != normalized]
    if len(kept) == len(employees):
        return False
    workbook["Employees"] = kept
    save_workbook_data(workbook)
    return True


def remove_employee_from_excel_data(email):
    if not EXCEL_PATH.exists():
        return False

    normalized = email.strip().lower()
    workbook = load_workbook(EXCEL_PATH)
    sheet = workbook["Employees"]
    headers = excel_headers(sheet)
    email_column = headers.index("Email") + 1

    for row_index in range(2, sheet.max_row + 1):
        if str(sheet.cell(row_index, email_column).value or "").strip().lower() == normalized:
            sheet.delete_rows(row_index, 1)
            workbook.save(EXCEL_PATH)
            workbook.close()
            return True

    workbook.close()
    return False


def is_admin(email):
    normalized = email.strip().lower()
    if normalized == ROOT_ADMIN_EMAIL:
        return True
    return bool(load_users().get(normalized, {}).get("is_admin"))


def set_admin_privilege(email, enabled, current_user_email):
    normalized = email.strip().lower()
    if not is_admin(current_user_email):
        return False, "Only an admin can change administration privileges."
    if normalized == ROOT_ADMIN_EMAIL and not enabled:
        return False, "Mohammed Alharbi's administration access cannot be removed."
    if normalized == current_user_email.strip().lower() and not enabled:
        return False, "You cannot remove your own administration access."
    if not normalized:
        return False, "Employee email is required."
    if normalized not in employee_emails() and normalized not in load_users():
        return False, "User was not found in the dashboard."

    users = load_users()
    employee = employee_by_email(normalized)
    record = users.get(normalized, {
        "email": normalized,
        "first_name": employee.get("Employee Name First Name", ""),
        "last_name": employee.get("Employee Name Last Name", ""),
        "created_at": int(time.time()),
    })
    record["is_admin"] = bool(enabled)
    record["approval_status"] = record.get("approval_status", "approved")
    users[normalized] = record
    save_users(users)
    return True, ""


def delete_employee(email, current_user_email):
    normalized = email.strip().lower()
    current = current_user_email.strip().lower()
    if not is_admin(current):
        return False, "Only an admin can delete users."
    if normalized == current:
        return False, "You cannot delete your own account."
    if not normalized:
        return False, "Employee email is required."

    removed_dashboard = remove_employee_from_dashboard_data(normalized)
    removed_excel = remove_employee_from_excel_data(normalized)
    users = load_users()
    removed_user = users.pop(normalized, None) is not None
    if removed_user:
        save_users(users)

    if not any([removed_dashboard, removed_excel, removed_user]):
        return False, "User was not found."
    return True, ""


def load_users():
    if not USERS_PATH.exists():
        return {}
    return json.loads(USERS_PATH.read_text(encoding="utf-8"))


def save_users(users):
    USERS_PATH.write_text(json.dumps(users, indent=2, sort_keys=True), encoding="utf-8")


def verify_or_create_user(email, password):
    normalized = email.strip().lower()
    users = load_users()
    if normalized not in employee_emails() and normalized not in users:
        return False, "Use your company email address.", ""

    record = users.get(normalized)
    if record and record.get("approval_status") == "pending":
        return False, "Your account is waiting for admin approval.", ""
    if record and record.get("approval_status") == "rejected":
        return False, "Your account request was not approved. Contact an administrator.", ""
    if record and record.get("password_hash"):
        stored_hash = record.get("password_hash", "").encode("utf-8")
        if bcrypt.checkpw(password.encode("utf-8"), stored_hash):
            return True, "", first_name_for_email(normalized)
        return False, "Invalid username or password.", ""

    employee = employee_by_email(normalized)
    password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    users[normalized] = {
        **(record or {}),
        "email": normalized,
        "first_name": (record or {}).get("first_name") or employee.get("Employee Name First Name", ""),
        "last_name": (record or {}).get("last_name") or employee.get("Employee Name Last Name", ""),
        "password_hash": password_hash,
        "created_at": (record or {}).get("created_at", int(time.time())),
        "approval_status": (record or {}).get("approval_status", "approved"),
    }
    save_users(users)
    return True, "", first_name_for_email(normalized)


def first_name_for_email(email):
    normalized = email.strip().lower()
    user = load_users().get(normalized, {})
    if user.get("first_name"):
        return user["first_name"]

    employee = employee_by_email(normalized)
    if employee.get("Employee Name First Name"):
        return employee["Employee Name First Name"]

    local_part = normalized.split("@", 1)[0]
    return local_part.split(".", 1)[0].title()


def create_signup_user(first_name, last_name, email, password, confirm_password):
    first_name = first_name.strip()
    last_name = last_name.strip()
    normalized = email.strip().lower()
    domain = company_domain()

    if not first_name or not last_name or not normalized:
        return False, "First name, last name, and company email are required."
    if not normalized.endswith(f"@{domain}"):
        return False, f"Use your @{domain} company email address."
    if len(password) < 8:
        return False, "Password must be at least 8 characters."
    if password != confirm_password:
        return False, "Passwords do not match."

    users = load_users()
    if normalized in users:
        return False, "A user with this email already exists."

    password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    users[normalized] = {
        "email": normalized,
        "first_name": first_name,
        "last_name": last_name,
        "password_hash": password_hash,
        "created_at": int(time.time()),
        "approval_status": "pending",
    }
    save_users(users)
    return True, ""


def pending_signup_users():
    users = load_users()
    return [
        {
            "email": email,
            "first_name": user.get("first_name", ""),
            "last_name": user.get("last_name", ""),
            "created_at": user.get("created_at", 0),
        }
        for email, user in sorted(users.items())
        if user.get("approval_status") == "pending"
    ]


def review_signup(email, action, current_user_email):
    normalized = email.strip().lower()
    action = action.strip().lower()
    if not is_admin(current_user_email):
        return False, "Only an admin can review signups.", {}

    users = load_users()
    user = users.get(normalized)
    if not user or user.get("approval_status") != "pending":
        return False, "Pending signup was not found.", {}

    if action == "approve":
        employee = add_employee_record(
            user.get("first_name", ""),
            user.get("last_name", ""),
            normalized,
        )
        user["approval_status"] = "approved"
        user["approved_at"] = int(time.time())
        user["approved_by"] = current_user_email.strip().lower()
        users[normalized] = user
        save_users(users)
        return True, "", {"employee": employee, "pending_users": pending_signup_users()}

    if action == "reject":
        users.pop(normalized, None)
        save_users(users)
        return True, "", {"pending_users": pending_signup_users()}

    return False, "Choose approve or reject.", {}


def validate_recovery_email(email):
    normalized = email.strip().lower()
    if not normalized:
        return False, "Email is required."
    if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", normalized):
        return False, "Enter a valid email address."
    return True, ""


def send_resend_email(to_email, subject, html, text):
    api_key = os.environ.get("RESEND_API_KEY")
    if not api_key:
        raise RuntimeError("RESEND_API_KEY is missing from .env")

    from_email = os.environ.get("RESEND_FROM_EMAIL", "Enterprise PMO <onboarding@resend.dev>")
    body = json.dumps({
        "from": from_email,
        "to": [to_email],
        "subject": subject,
        "html": html,
        "text": text,
    }).encode("utf-8")
    request = Request(
        "https://api.resend.com/emails",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "User-Agent": "enterprise-pmo-dashboard/1.0",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=30) as response:
            return json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        if error.code == 403 and "can only send testing emails" in detail:
            raise RuntimeError(
                "Resend is blocking this recipient because onboarding@resend.dev is a test sender. "
                "Send to your Resend account email, or verify a domain in Resend and set RESEND_FROM_EMAIL to that domain."
            ) from error
        if error.code == 403 and "domain is not verified" in detail:
            raise RuntimeError(
                "Resend is blocking this sender because the RESEND_FROM_EMAIL domain is not verified."
            ) from error
        raise RuntimeError(f"Resend API error {error.code}: {detail}") from error
    except URLError as error:
        raise RuntimeError(f"Could not reach Resend API: {error.reason}") from error


def request_password_reset(email):
    normalized = email.strip().lower()
    ok, error = validate_recovery_email(normalized)
    if not ok:
        return False, error

    code = f"{secrets.randbelow(1000000):06d}"
    PASSWORD_RESETS[normalized] = {
        "code": code,
        "expires_at": time.time() + RESET_CODE_SECONDS,
        "attempts": 0,
    }
    send_resend_email(
        normalized,
        "Enterprise PMO password reset code",
        f"""
        <div style="font-family:Arial,sans-serif;line-height:1.5;color:#172026">
          <h2>Password reset code</h2>
          <p>Your Enterprise PMO Dashboard verification code is:</p>
          <p style="font-size:28px;font-weight:700;letter-spacing:6px">{code}</p>
          <p>This code expires in 10 minutes. If you did not request it, you can ignore this email.</p>
        </div>
        """,
        f"Your Enterprise PMO Dashboard verification code is {code}. It expires in 10 minutes.",
    )
    return True, ""


def verify_password_reset_code(email, code):
    normalized = email.strip().lower()
    record = PASSWORD_RESETS.get(normalized)
    if not record:
        return False, "Request a new verification code first.", ""
    if record["expires_at"] < time.time():
        PASSWORD_RESETS.pop(normalized, None)
        return False, "Verification code expired. Request a new code.", ""
    if record.get("attempts", 0) >= 5:
        PASSWORD_RESETS.pop(normalized, None)
        return False, "Too many incorrect attempts. Request a new code.", ""

    record["attempts"] = record.get("attempts", 0) + 1
    if not secrets.compare_digest(str(code).strip(), record["code"]):
        return False, "Invalid verification code.", ""

    token = secrets.token_urlsafe(32)
    record["token"] = token
    record["token_expires_at"] = time.time() + RESET_TOKEN_SECONDS
    record.pop("code", None)
    return True, "", token


def reset_password(token, password, confirm_password):
    if len(password) < 8:
        return False, "Password must be at least 8 characters."
    if password != confirm_password:
        return False, "Passwords do not match."

    normalized = ""
    reset_record = {}
    for email, record in PASSWORD_RESETS.items():
        if secrets.compare_digest(str(record.get("token", "")), token):
            normalized = email
            reset_record = record
            break

    if not normalized or reset_record.get("token_expires_at", 0) < time.time():
        if normalized:
            PASSWORD_RESETS.pop(normalized, None)
        return False, "Reset session expired. Request a new code."

    employee = employee_by_email(normalized)
    users = load_users()
    existing = users.get(normalized, {})
    password_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")
    users[normalized] = {
        **existing,
        "email": normalized,
        "first_name": existing.get("first_name") or employee.get("Employee Name First Name", ""),
        "last_name": existing.get("last_name") or employee.get("Employee Name Last Name", ""),
        "password_hash": password_hash,
        "created_at": existing.get("created_at", int(time.time())),
        "password_reset_at": int(time.time()),
    }
    save_users(users)
    PASSWORD_RESETS.pop(normalized, None)
    return True, ""


def parse_cookies(header):
    cookies = {}
    for part in (header or "").split(";"):
        if "=" in part:
            key, value = part.strip().split("=", 1)
            cookies[key] = value
    return cookies


def session_email(headers):
    token = parse_cookies(headers.get("Cookie")).get("dashboard_session")
    if not token:
        return ""
    session = SESSIONS.get(token)
    if not session or session["expires_at"] < time.time():
        SESSIONS.pop(token, None)
        return ""
    return session["email"]


def create_session(email):
    token = secrets.token_urlsafe(32)
    SESSIONS[token] = {
        "email": email,
        "expires_at": time.time() + SESSION_SECONDS,
    }
    return token


def require_admin(headers):
    email = session_email(headers)
    if not email:
        return "", "Please log in first."
    if not is_admin(email):
        return email, "Only an admin can perform this action."
    return email, ""


def admin_emails():
    emails = {ROOT_ADMIN_EMAIL}
    users = load_users()
    for email, user in users.items():
        if user.get("is_admin"):
            emails.add(email)
    return sorted(emails)


def response_text(payload):
    if isinstance(payload.get("output_text"), str):
        return payload["output_text"].strip()

    parts = []
    for item in payload.get("output", []):
        for content in item.get("content", []):
            if content.get("type") in {"output_text", "text"} and isinstance(content.get("text"), str):
                parts.append(content["text"])
    return "\n".join(parts).strip()


def ask_openai(question, context, history):
    api_key = os.environ.get("OPENAI_API_KEY")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is missing from .env")

    model = os.environ.get("OPENAI_MODEL", "gpt-5.5")
    system_prompt = (
        "You are the Enterprise PMO Dashboard Assistant. Answer using only the "
        "dashboard context provided by the app. If the answer is not in that "
        "context, say you do not have enough dashboard data to answer. Keep "
        "answers concise and include exact counts or SAR values when available."
    )
    user_payload = {
        "question": question,
        "dashboard_context": context,
        "recent_chat_history": history[-8:],
    }
    body = json.dumps({
        "model": model,
        "input": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
        ],
    }).encode("utf-8")

    request = Request(
        "https://api.openai.com/v1/responses",
        data=body,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )

    try:
        with urlopen(request, timeout=45) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except HTTPError as error:
        detail = error.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"OpenAI API error {error.code}: {detail}") from error
    except URLError as error:
        raise RuntimeError(f"Could not reach OpenAI API: {error.reason}") from error

    answer = response_text(payload)
    if not answer:
        raise RuntimeError("OpenAI returned an empty response")
    return answer


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=str(ROOT), **kwargs)

    def end_headers(self):
        if self.path.endswith((".html", ".js", ".css")) or self.path in {"/", "/index.html"}:
            self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def send_json(self, status, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def send_json_with_cookie(self, status, payload, cookie):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Set-Cookie", cookie)
        self.end_headers()
        self.wfile.write(body)

    def authenticated(self):
        return bool(session_email(self.headers))

    def do_GET(self):
        if self.path == "/api/me":
            email = session_email(self.headers)
            if not email:
                self.send_json(401, {"error": "Please log in first."})
                return
            employee = employee_by_email(email)
            user_is_admin = is_admin(email)
            self.send_json(200, {
                "email": email,
                "first_name": first_name_for_email(email),
                "last_name": load_users().get(email, {}).get("last_name") or employee.get("Employee Name Last Name", ""),
                "can_delete_users": user_is_admin,
                "is_admin": user_is_admin,
                "admin_emails": admin_emails() if user_is_admin else [],
                "pending_users": pending_signup_users() if user_is_admin else [],
            })
            return

        if self.path in {"/", "/index.html"} and not self.authenticated():
            login_path = ROOT / "login.html"
            body = login_path.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if self.path in {"/app.js", "/data.js"} and not self.authenticated():
            self.send_error(401)
            return

        if self.path.startswith("/api/"):
            self.send_json(404, {"error": "API route was not found."})
            return

        super().do_GET()

    def do_POST(self):
        if self.path == "/api/login":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                email = str(payload.get("email", "")).strip()
                password = str(payload.get("password", ""))
                if not email or not password:
                    self.send_json(400, {"error": "Email and password are required."})
                    return

                ok, error, first_name = verify_or_create_user(email, password)
                if not ok:
                    self.send_json(401, {"error": error})
                    return

                normalized = email.lower()
                token = create_session(normalized)
                cookie = f"dashboard_session={token}; HttpOnly; SameSite=Lax; Path=/; Max-Age={SESSION_SECONDS}"
                user_is_admin = is_admin(normalized)
                self.send_json_with_cookie(200, {
                    "email": normalized,
                    "first_name": first_name,
                    "can_delete_users": user_is_admin,
                    "is_admin": user_is_admin,
                    "admin_emails": admin_emails() if user_is_admin else [],
                    "pending_users": pending_signup_users() if user_is_admin else [],
                }, cookie)
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/signup":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, error = create_signup_user(
                    str(payload.get("first_name", "")),
                    str(payload.get("last_name", "")),
                    str(payload.get("email", "")),
                    str(payload.get("password", "")),
                    str(payload.get("confirm_password", "")),
                )
                if not ok:
                    self.send_json(400, {"error": error})
                    return
                self.send_json(201, {"ok": True})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/forgot-password":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, error = request_password_reset(str(payload.get("email", "")))
                if not ok:
                    self.send_json(400, {"error": error})
                    return
                self.send_json(200, {"ok": True})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/verify-reset-code":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, error, token = verify_password_reset_code(
                    str(payload.get("email", "")),
                    str(payload.get("code", "")),
                )
                if not ok:
                    self.send_json(400, {"error": error})
                    return
                self.send_json(200, {"ok": True, "reset_token": token})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/reset-password":
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, error = reset_password(
                    str(payload.get("reset_token", "")),
                    str(payload.get("password", "")),
                    str(payload.get("confirm_password", "")),
                )
                if not ok:
                    self.send_json(400, {"error": error})
                    return
                self.send_json(200, {"ok": True})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/logout":
            token = parse_cookies(self.headers.get("Cookie")).get("dashboard_session")
            if token:
                SESSIONS.pop(token, None)
            cookie = "dashboard_session=; HttpOnly; SameSite=Lax; Path=/; Max-Age=0"
            self.send_json_with_cookie(200, {"ok": True}, cookie)
            return

        if self.path == "/api/delete-user":
            current_user = session_email(self.headers)
            if not current_user:
                self.send_json(401, {"error": "Please log in first."})
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, error = delete_employee(str(payload.get("email", "")), current_user)
                if not ok:
                    self.send_json(403, {"error": error})
                    return
                self.send_json(200, {"ok": True})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/add-user":
            current_user, error = require_admin(self.headers)
            if error:
                self.send_json(401 if not current_user else 403, {"error": error})
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, error, employee = create_dashboard_employee(
                    str(payload.get("first_name", "")),
                    str(payload.get("last_name", "")),
                    str(payload.get("email", "")),
                )
                if not ok:
                    self.send_json(400, {"error": error})
                    return
                self.send_json(201, {"ok": True, "employee": employee})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/review-signup":
            current_user, error = require_admin(self.headers)
            if error:
                self.send_json(401 if not current_user else 403, {"error": error})
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, review_error, result = review_signup(
                    str(payload.get("email", "")),
                    str(payload.get("action", "")),
                    current_user,
                )
                if not ok:
                    self.send_json(400, {"error": review_error})
                    return
                self.send_json(200, {"ok": True, **result})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/update-workbook":
            current_user, error = require_admin(self.headers)
            if error:
                self.send_json(401 if not current_user else 403, {"error": error})
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, update_error = replace_workbook_data(payload.get("workbook", {}))
                if not ok:
                    self.send_json(400, {"error": update_error})
                    return
                self.send_json(200, {"ok": True})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path == "/api/set-admin":
            current_user, error = require_admin(self.headers)
            if error:
                self.send_json(401 if not current_user else 403, {"error": error})
                return
            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                ok, privilege_error = set_admin_privilege(
                    str(payload.get("email", "")),
                    bool(payload.get("is_admin")),
                    current_user,
                )
                if not ok:
                    self.send_json(400, {"error": privilege_error})
                    return
                self.send_json(200, {"ok": True, "admin_emails": admin_emails()})
            except Exception as error:
                self.send_json(500, {"error": str(error)})
            return

        if self.path != "/api/chat":
            if self.path.startswith("/api/"):
                self.send_json(404, {"error": "API route was not found."})
                return
            self.send_error(404)
            return

        if not self.authenticated():
            self.send_json(401, {"error": "Please log in first."})
            return

        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8"))
            question = str(payload.get("question", "")).strip()
            if not question:
                self.send_json(400, {"error": "Question is required."})
                return

            answer = ask_openai(
                question,
                payload.get("context", {}),
                payload.get("history", []),
            )
            self.send_json(200, {"answer": answer})
        except Exception as error:
            self.send_json(500, {"error": str(error)})


if __name__ == "__main__":
    load_env()
    sync_registered_users_to_employees()
    port = int(os.environ.get("PORT", "8000"))
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    print(f"Dashboard server running at http://127.0.0.1:{port}")
    server.serve_forever()
